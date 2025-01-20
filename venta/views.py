from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import *
from .models import *
from datetime import datetime
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, letter
import json
import openpyxl
from openpyxl.styles import Font
from django.db import transaction

# Create your views here.
def rol_requerido(rol_permitido):
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            usuario_rol = request.session['usuario_rol']
            if usuario_rol == rol_permitido:
                print('acceso permitido')
                return view_func(request, *args, **kwargs)
            print('acceso denegado')
            return redirect('home')
        return _wrapped_view
    return decorator

@login_required
def home(request):
    ventas = Venta.objects.all().count()
    clientes = Cliente.objects.all().count()
    negocios = Negocio.objects.all().count()
    marcas = Marca.objects.all().count()
    usuario = Usuario.objects.get(user=request.user)
    request.session['usuario_rol']=usuario.rol
    print(request.session['usuario_rol'])
    context = {
        'ventas': ventas,
        'clientes': clientes,
        'negocios': negocios,
        'marcas': marcas
    }
    return render(request, 'index.html', context)

@login_required
@rol_requerido(1)
def negocio(request):
    negocios = Negocio.objects.all()
    context = {'negocios': negocios}
    return render(request, 'negocio/list.html', context)

@login_required
@rol_requerido(1)
def negocio_store(request):
    nf = NegocioForm(request.POST or None)
    cf = ContactoForm(request.POST or None)
    ff = FiscalForm(request.POST or None)
    df = DomicilioForm(request.POST or None)
    context = {
        'nf': nf,
        'cf': cf,
        'ff': ff,
        'df': df
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                print('transaction begin')
                if nf.is_valid() and cf.is_valid() and ff.is_valid() and df.is_valid():
                    n = nf.save(commit=False)
                    f = ff.save(commit=False)
                    d = df.save()
                    c = cf.save()
                    f.domicilio = d
                    f.save()
                    n.contacto = c
                    n.fiscal = f
                    n.save()
                    print('transaction end')

                    return redirect('negocios')
                else:
                    print(nf.errors)
                    print(cf.errors)
                    print(ff.errors)
                    print(df.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'negocio/crear.html', context)

@login_required
@rol_requerido(1)
def negocio_update(request, pk):
    negocio = Negocio.objects.get(pk=pk)
    contacto = negocio.contacto
    fiscal = negocio.fiscal
    domicilio = fiscal.domicilio
    nf = NegocioForm(request.POST or None, instance=negocio)
    cf = ContactoForm(request.POST or None, instance=contacto)
    ff = FiscalForm(request.POST or None, instance=fiscal)
    df = DomicilioForm(request.POST or None, instance=domicilio)
    context = {
        'nf': nf,
        'cf': cf,
        'ff': ff,
        'df': df
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if nf.is_valid() and cf.is_valid() and ff.is_valid() and df.is_valid():
                    df.save()
                    ff.save()
                    cf.save()
                    nf.save()

                    return redirect('negocios')
                else:
                    print(nf.errors)
                    print(cf.errors)
                    print(ff.errors)
                    print(df.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'negocio/editar.html', context)

@login_required
@rol_requerido(1)
def negocio_destroy(request, pk):
    negocio = Negocio.objects.get(pk=pk)
    contacto = negocio.contacto
    fiscal = negocio.fiscal
    domicilio = fiscal.domicilio

    if request.method == 'POST':
        with transaction.atomic():
            domicilio.delete()
            fiscal.delete()
            contacto.delete()
            negocio.delete()

            return redirect('negocios')

@login_required
def marca(request):
    marcas = Marca.objects.all()
    context = {'marcas': marcas}
    return render(request, 'marca/list.html', context)

def marca_store(request):
    form = MarcaForm(request.POST or None)
    context = {'form': form}
    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('marcas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'marca/crear.html', context)

@login_required
def marca_update(request, pk):
    marca = Marca.objects.get(pk=pk)
    form = MarcaForm(request.POST or None, instance=marca)
    context = {'form': form}
    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('marcas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'marca/editar.html', context)

@login_required
def marca_destroy(request, pk):
    marca = Marca.objects.get(pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            marca.delete()
            return redirect('marcas')

@login_required
@rol_requerido(1)
def cliente(request):
    clientes = Cliente.objects.all()
    context = {'clientes': clientes}

    return render(request, 'cliente/list.html', context)

@login_required
@rol_requerido(1)
def cliente_store(request):
    clif = ClienteForm(request.POST or None)
    cf = ContactoForm(request.POST or None)
    ff = FiscalForm(request.POST or None)
    df = DomicilioForm(request.POST or None)
    context = {
        'clif': clif,
        'cf': cf,
        'ff': ff,
        'df': df
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if clif.is_valid() and cf.is_valid() and ff.is_valid() and df.is_valid():
                    cl = clif.save(commit=False)
                    f = ff.save(commit=False)
                    d = df.save()
                    c = cf.save()
                    f.domicilio = d
                    f.save()
                    cl.contacto = c
                    cl.fiscal = f
                    cl.save()

                    return redirect('clientes')
                else:
                    print(clif.errors)
                    print(cf.errors)
                    print(ff.errors)
                    print(df.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'cliente/crear.html', context)

@login_required
@rol_requerido(1)
def cliente_update(request, pk):
    cliente = Cliente.objects.get(pk=pk)
    contacto = cliente.contacto
    fiscal = cliente.fiscal
    domicilio = fiscal.domicilio
    clif = ClienteForm(request.POST or None, instance=cliente)
    cf = ContactoForm(request.POST or None, instance=contacto)
    ff = FiscalForm(request.POST or None, instance=fiscal)
    df = DomicilioForm(request.POST or None, instance=domicilio)
    context = {
        'clif': clif,
        'cf': cf,
        'ff': ff,
        'df': df
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if clif.is_valid() and cf.is_valid() and ff.is_valid() and df.is_valid():
                    df.save()
                    ff.save()
                    cf.save()
                    clif.save()

                    return redirect('clientes')
                else:
                    print(clif.errors)
                    print(cf.errors)
                    print(ff.errors)
                    print(df.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'cliente/editar.html', context)

@login_required
@rol_requerido(1)
def cliente_destroy(request, pk):
    cliente = Cliente.objects.get(pk=pk)
    contacto = cliente.contacto
    fiscal = cliente.fiscal
    domicilio = fiscal.domicilio

    if request.method == 'POST':
        with transaction.atomic():
            domicilio.delete()
            fiscal.delete()
            contacto.delete()
            cliente.delete()

        return redirect('clientes')

@login_required
def caja(request):
    cajas = Caja.objects.all()
    usuario = Usuario.objects.get(user=request.user)
    context = {
        'cajas': cajas,
        'usuario': usuario,
    }

    return render(request, 'caja/list.html', context)

@login_required
def caja_store(request):
    form = CajaForm(request.POST or None)
    context = {'form':form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('cajas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/crear.html', context)

@login_required
def caja_update(request, pk):
    caja = Caja.objects.get(pk=pk)
    form = CajaForm(request.POST or None, instance=caja)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('cajas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/editar.html', context)

@login_required
def caja_destroy(request, pk):
    caja = Caja.objects.get(pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            caja.delete()
            return redirect('cajas')

@login_required
def departamento(request):
    departamentos = Departamento.objects.all()
    context = {'departamentos': departamentos}

    return render(request, 'departamento/list.html', context)

@login_required
def departamento_store(request):
    form = DepartamentoForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('departamentos')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'departamento/crear.html', context)

@login_required
def departamento_update(request, pk):
    departamento = Departamento.objects.get(pk=pk)
    form = DepartamentoForm(request.POST or None, instance=departamento)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('departamentos')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'departamento/editar.html', context)

@login_required
def departamento_destroy(request, pk):
    departamento = Departamento.objects.get(pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            departamento.delete()

            return redirect('departamentos')

@login_required
def medida(request):
    medidas = Medida.objects.all()
    context = {'medidas': medidas}

    return render(request, 'medida/list.html', context)

@login_required
def medida_store(request):
    form = MedidaForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('medidas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'medida/crear.html', context)

@login_required
def medida_update(request, pk):
    medida = Medida.objects.get(pk=pk)
    form = MedidaForm(request.POST or None, instance=medida)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('medidas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'medida/editar.html', context)

@login_required
def medida_destroy(request, pk):
    medida = Medida.objects.get(pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            medida.delete()
            return redirect('medidas')

@login_required
def producto(request):
    productos = Producto.objects.all()
    context = {'productos': productos}

    return render(request, 'producto/list.html', context)

@login_required
def producto_store(request):
    form = ProductoForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('productos')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'producto/crear.html', context)

@login_required
def producto_update(request, pk):
    producto = Producto.objects.get(pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    context = {'form': form}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    form.save()
                    return redirect('productos')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'producto/editar.html', context)

@login_required
def producto_destroy(request, pk):
    producto = Producto.objects.get(pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            producto.delete()
            return redirect('productos')

@login_required
@rol_requerido(1)
def usuario(request):
    usuarios = Usuario.objects.all()
    context = {'usuarios': usuarios}

    return render(request, 'usuario/list.html', context)

@login_required
@rol_requerido(1)
def usuario_store(request):
    ef = EmpleadoForm(request.POST or None, request.FILES or None)
    uf = UserForm(request.POST or None)
    cf = ContactoForm(request.POST or None)
    context = {
        'ef': ef,
        'uf': uf,
        'cf': cf
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if ef.is_valid() and uf.is_valid() and cf.is_valid():
                    e = ef.save(commit=False)
                    u = uf.save()
                    c = cf.save()
                    e.user = u
                    e.contacto = c
                    e.save()
                    return redirect('usuarios')
                else:
                    print(ef.errors)
                    print(uf.errors)
                    print(cf.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'usuario/crear.html', context)

@login_required
@rol_requerido(1)
def usuario_update(request, pk):
    usuario = Usuario.objects.get(pk=pk)
    user = usuario.user
    contacto = usuario.contacto
    ef = EmpleadoForm(request.POST or None, request.FILES or None, instance=usuario)
    uf = UserForm(request.POST or None, instance=user)
    cf = ContactoForm(request.POST or None, instance=contacto)
    context = {
        'ef': ef,
        'uf': uf,
        'cf': cf
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if ef.is_valid() and uf.is_valid() and cf.is_valid():
                    ef.save(commit=False)
                    uf.save()
                    cf.save()
                    return redirect('usuarios')
                else:
                    print(ef.errors)
                    print(uf.errors)
                    print(cf.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'usuario/editar.html', context)

@login_required
@rol_requerido(1)
def usuario_destroy(request, pk):
    usuario = Usuario.objects.get(pk=pk)
    user = usuario.user
    contacto = usuario.contacto

    if request.method == 'POST':
        with transaction.atomic():
            user.delete()
            contacto.delete()
            usuario.delete()

            return redirect('usuarios')

def singing(request):
    if request.method == 'POST':
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        login(request, user)
        return redirect('home')

    return render(request, 'login.html')

@login_required
def signout(request):
    request.session['usuario_rol']=0
    logout(request)
    return redirect('login')

def singup(request):
    uf = UserForm(request.POST or None)
    rf = RegistroForm(request.POST or None)
    context = {
        'uf': uf,
        'rf': rf
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if uf.is_valid() and rf.is_valid():
                    r = rf.save(commit=False)
                    u = uf.save()
                    c = Contacto.objects.create()
                    r.user = u
                    r.contacto = c
                    r.rol = 2
                    r.save()
                    login(request, u)
                    return redirect('home')
                else:
                    print(uf.errors)
                    print(rf.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'register.html', context)

@login_required
def perfil(request):
    usuario = Usuario.objects.get(user=request.user)
    user = usuario.user
    contacto = usuario.contacto
    ef = PerfilForm(request.POST or None, request.FILES or None, instance=usuario)
    uf = UserForm(request.POST or None, instance=user)
    cf = ContactoForm(request.POST or None, instance=contacto)
    context = {
        'ef': ef,
        'uf': uf,
        'cf': cf
    }

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if ef.is_valid() and uf.is_valid() and cf.is_valid():
                    ef.save(commit=False)
                    uf.save()
                    cf.save()
                    return redirect('usuarios')
                else:
                    print(ef.errors)
                    print(uf.errors)
                    print(cf.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'perfil.html', context)

@login_required
def caja_open(request, pk):
    caja = Caja.objects.get(pk=pk)
    usuario = Usuario.objects.get(user=request.user)
    form = AbrirCajaForm(request.POST or None)
    context = {'form': form}
    request.session['caja_abierta'] = 0

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if form.is_valid():
                    f = form.save(commit=False)
                    f.caja = caja
                    f.movimento = 'A'
                    f.empleado = usuario
                    f.fecha = datetime.now()
                    request.session['caja_abierta'] = caja.id
                    f.save()
                    return redirect('cajas')
                else:
                    print(form.errors)
        except Exception as e:
            print(f'error: {e}')

    return render(request, 'caja/abrir.html', context)

@login_required
def caja_close(request, pk):
    caja = Caja.objects.get(pk=pk)
    caja_abierta = MCaja.objects.filter(caja=caja, movimento='A').order_by('-fecha').first()   
    empleado = Usuario.objects.get(user=request.user)

    if request.method == 'POST':
        with transaction.atomic():
            MCaja.objects.create(
                fecha=datetime.now(),
                movimento='C',
                caja_id=caja.id,
                empleado_id=empleado.id,
                ma=caja_abierta.ma,
                mc=request.session['monto'],
                ganacia=float(request.session['monto'])-float(caja_abierta.ma)
            )
            request.session['caja_abierta'] = 0
            request.session['monto']=0
            return redirect('cajas')

@login_required
def venta(request):
    usuario = Usuario.objects.get(user=request.user)
    ventas = ''
    if usuario.rol==1:
        ventas = Venta.objects.all()
    else:
        ventas = Venta.objects.filter(empleado=usuario)
    context = {'ventas': ventas}

    return render(request, 'venta/list.html', context)

@login_required
def venta_store(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        productos = data.get('producto', [])
        print(data)
        request.session['monto']=0
        try:
            with transaction.atomic():
                venta = Venta.objects.create(
                    fecha=datetime.now(),
                    caja_id=request.session['caja_abierta'],
                    empleado_id=Usuario.objects.get(user=request.user).id,
                    cliente_id=data.get('cliente', {}).get('id', 0),
                    total=data.get('subtotal', 0),
                    total_descuento=data.get('total', 0),
                    importe=data.get('importe', 0),
                    cambio=float(data.get('importe', 0))-float(data.get('total', 0)),
                    mp=data.get('metodo', '')
                )
                request.session['monto']+=float(data.get('total', 0))
                for producto in productos:
                    pro = Producto.objects.get(pk=producto.get('id', 0))
                    total_producto = float(producto.get('cantidad', 0))*float(pro.precio)
                    producto_id = producto.get('id', 0)
                    cantidad = producto.get('cantidad', 0)
                    DetalleVenta.objects.create(
                        venta_id=venta.id,
                        producto_id=producto_id,
                        cantidad=cantidad,
                        importe=data.get('importe', 0),
                        descuento=data.get('descuento', 0),
                        total=total_producto
                    )
                return JsonResponse({'success': True, 'message': 'Mensaje enviado correctamente', 'id': venta.id})
        except Exception as e:
            print(f'error: {e}')
            return JsonResponse({'success': False, 'message': e})
    return render(request, 'venta/crear.html')

def autocomplete_producto(request):
    if 'term' in request.GET:
        qs = Producto.objects.filter(nombre__icontains=request.GET.get("term"))
        titles = list()
        for producto in qs:
            datos = {
                'id': producto.id,
                'label': producto.nombre,
                'precio': producto.precio,
                'existencia': producto.existencia
            }
            titles.append(datos)
        return JsonResponse(titles, safe=False)

def autocomplete_cliente(request):
    if 'term' in request.GET:
        qs = Cliente.objects.filter(nombres__icontains=request.GET.get("term"))
        titles = list()
        for cliente in qs:
            datos = {
            'id': cliente.id,
            'label': cliente.nombres
            }
            titles.append(datos)
        return JsonResponse(titles, safe=False)

@login_required
def ticket(request, pk):
    venta = Venta.objects.get(pk=pk)
    detalles = DetalleVenta.objects.filter(venta=venta)
    
    # Tamaño personalizado para el ticket (8 cm x 20 cm en puntos)
    ticket_width = 227  # 8 cm
    ticket_height = 567  # 20 cm
    ticket_size = (ticket_width, ticket_height)

    # Crear un buffer en memoria
    buffer = BytesIO()
    
    # Crear canvas en el buffer con tamaño de ticket
    c = canvas.Canvas(buffer, pagesize=ticket_size)

    # Margen inicial
    x_inicial = 10  # Reducir margen para el ticket
    y_actual = ticket_height - 20  # Comenzar desde el tope

    # Encabezado
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x_inicial, y_actual, "Punto de Venta")
    y_actual -= 12

    c.setFont("Helvetica", 8)
    c.drawString(x_inicial, y_actual, f"RFC: SUL-010720-JN8")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, f"Fecha: {venta.fecha.strftime('%d/%m/%Y - %H:%M:%S')}")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, "Ave. Lopez Mateos #1590-C Playa Ensenada")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, "Ensenada, B.C. - TEL. (646) 152-1901")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, "FAX. (646) 152-1901")
    y_actual -= 20

    # Información del vendedor
    c.drawString(x_inicial, y_actual, f"Vendedor: {venta.empleado.nombres if venta.empleado else 'No asignado'}")
    y_actual -= 20

    # Detalle de los productos
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x_inicial, y_actual, "CANT")
    c.drawString(x_inicial + 50, y_actual, "DESCRIPCION")
    c.drawString(x_inicial + 150, y_actual, "PRECIO")
    y_actual -= 15

    c.setFont("Helvetica", 10)
    for detalle in detalles:
        c.drawString(x_inicial, y_actual, str(detalle.cantidad))
        c.drawString(x_inicial + 50, y_actual, detalle.producto.nombre)
        c.drawString(x_inicial + 150, y_actual, f"${detalle.producto.precio:.2f}")
        y_actual -= 15

    y_actual -= 15

    # Total y detalles del pago
    c.drawString(x_inicial, y_actual, f'Subtotal: ${venta.total:.2f}')
    y_actual -= 12
    c.drawString(x_inicial, y_actual, f"Total: ${venta.total_descuento:.2f}")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, f"Importe: ${venta.importe:.2f}")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, f"Cambio: ${venta.cambio:.2f}")
    y_actual -= 20

    # Pie de página
    c.drawString(x_inicial, y_actual, "Gracias por su compra!")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, f"Computadora: CWIN-SYSTEMS")
    y_actual -= 12
    c.drawString(x_inicial, y_actual, f"Folio Interno: {venta.id}")

    # Guardar el PDF en el buffer
    c.save()

    # Mover el cursor del buffer al inicio
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{venta.id}.pdf"'
    return response

@login_required
def factura(request, pk):
    venta = Venta.objects.get(pk=pk)
    detalles = DetalleVenta.objects.filter(venta=venta)
    descuento = float(venta.total)-float(venta.total_descuento)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="factura_{venta.id}.pdf"'

    # Crear el canvas de ReportLab
    pdf = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Encabezado de la factura
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, height - 50, "Factura de Venta")

    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, height - 80, f"Fecha: {venta.fecha.strftime('%d/%m/%Y %H:%M:%S')}")
    pdf.drawString(50, height - 100, f"Caja: {venta.caja.id if venta.caja else 'N/A'}")
    pdf.drawString(50, height - 120, f"Empleado: {venta.empleado.nombres if venta.empleado else 'N/A'}")
    pdf.drawString(50, height - 140, f"Cliente: {venta.cliente.nombres if venta.cliente else 'N/A'}")

    pdf.drawString(50, height - 160, f"Método de Pago: {venta.metodo()}")
    pdf.drawString(50, height - 180, f"Subtotal: ${venta.total:.2f}")
    pdf.drawString(50, height - 200, f"Descuento: ${descuento:.2f}")
    pdf.drawString(50, height - 220, f"Total a Pagar: ${venta.total_descuento:.2f}")

    # Dibujar encabezados de tabla
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, height - 240, "Producto")
    pdf.drawString(200, height - 240, "Cantidad")
    pdf.drawString(300, height - 240, "Precio Unitario")
    pdf.drawString(500, height - 240, "Total")

    # Dibujar los detalles de la venta
    pdf.setFont("Helvetica", 10)
    y = height - 260
    for detalle in detalles:
        total = float(detalle.cantidad)*float(detalle.producto.precio)
        pdf.drawString(50, y, detalle.producto.nombre if detalle.producto else "N/A")
        pdf.drawString(200, y, str(detalle.cantidad))
        pdf.drawString(300, y, f"${detalle.producto.precio}")
        pdf.drawString(500, y, f"${total:.2f}")
        y -= 20

        # Salto de página si el contenido es muy largo
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            y = height - 50

    # Finalizar el PDF
    pdf.save()
    return response

@login_required
@rol_requerido(1)
def excel(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Ventas'

    # Establecer el encabezado
    encabezados = [
        'ID', 'Fecha', 'Caja', 'Empleado', 'Cliente', 
        'SubTotal', 'Total', 'Importe', 'Cambio', 
        'Metodo de Pago'
    ]
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = sheet.cell(row=1, column=col_num, value=encabezado)
        celda.font = Font(bold=True)  # Hacer el encabezado en negrita

    # Obtener los datos del modelo
    ventas = Venta.objects.all()
    for row_num, venta in enumerate(ventas, start=2):
        sheet.cell(row=row_num, column=1, value=venta.id)
        sheet.cell(row=row_num, column=2, value=venta.fecha.strftime('%Y-%m-%d %H:%M:%S') if venta.fecha else '')
        sheet.cell(row=row_num, column=3, value=venta.caja.nombre if venta.caja else 'N/A')
        sheet.cell(row=row_num, column=4, value=venta.empleado.nombres if venta.empleado else 'N/A')
        sheet.cell(row=row_num, column=5, value=venta.cliente.nombres if venta.cliente else 'N/A')
        sheet.cell(row=row_num, column=6, value=venta.total)
        sheet.cell(row=row_num, column=7, value=venta.total_descuento)
        sheet.cell(row=row_num, column=8, value=venta.importe)
        sheet.cell(row=row_num, column=9, value=venta.cambio)
        sheet.cell(row=row_num, column=10, value=venta.metodo())

    # Configurar la respuesta HTTP para la descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="ventas.xlsx"'

    # Guardar el archivo Excel en la respuesta
    workbook.save(response)
    return response

@login_required
@rol_requerido(1)
def reporte(request):
    ventas = ''
    cajas = Caja.objects.all()
    usuarios = Usuario.objects.all()
    if request.method == 'POST':
        data = json.loads(request.body)
        filtros = {}
        formato = "%Y-%m-%d %H:%M:%S"
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        empleado = int(data.get('empleado'))
        caja = int(data.get('caja'))
        metodo = data.get('metodo')

        if fecha_inicio!='' and fecha_fin!='':
            fecha_formato = datetime.strptime(f"{fecha_inicio} 00:00:00", formato)
            fecha_formato_2 = datetime.strptime(f"{fecha_fin} 00:00:00", formato)
            filtros['fecha__range']=(fecha_formato, fecha_formato_2)

        if empleado!=0:
            filtros['empleado_id']=empleado

        if caja !=0:
            filtros['caja_id']=caja

        if metodo!='':
            filtros['mp']=metodo

        try:
            ventas = Venta.objects.filter(**filtros)
            ventas_data = []
            for v in ventas:
                ventas_data.append({
                    'id': v.id,
                    'fecha': v.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                    'caja': v.caja.nombre,
                    'empleado': v.empleado.nombres,
                    'cliente': v.cliente.nombres,
                    'total': float(v.total),
                    'total_descuento': float(v.total_descuento),
                    'importe': float(v.importe),
                    'cambio': float(v.cambio),
                    'metodo_pago': v.metodo()
                })
            return JsonResponse({'success': True, 'ventas': ventas_data})
        except Exception as e:
            print(f'error: {e}')
    else:
        ventas = Venta.objects.all()
    context = {
        'ventas': ventas,
        'cajas': cajas,
        'usuarios': usuarios
    }
    return render(request, 'venta/reporte.html', context)
